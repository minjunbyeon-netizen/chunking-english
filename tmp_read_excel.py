import openpyxl
import json
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\work\chunking-english\asset\final.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

# Read all rows
all_rows = []
for row in ws.iter_rows(values_only=True):
    cells = [str(c).strip() if c is not None else '' for c in row]
    all_rows.append(cells)

wb.close()

# Skip header row
data_rows = all_rows[1:]

# Build mapping: slug -> korean
mapping = {}  # { "have_a_dream": "가지다 꿈을", ... }
sentence_mapping = {}  # { "I_have_a_dream": "나는 가지다 꿈을", ... }

i = 0
while i < len(data_rows) - 1:
    en_row = data_rows[i]
    kr_row = data_rows[i + 1]
    
    # Skip empty rows
    if not en_row[0] and not en_row[1]:
        i += 1
        continue
    
    # Column 0 = sentence (I have a dream.)
    sentence_en = en_row[0].rstrip('.')
    sentence_kr = kr_row[0] if kr_row[0] else ''
    if sentence_en:
        slug = sentence_en.replace(' ', '_')
        sentence_mapping[slug] = sentence_kr
    
    # Columns 2-8 = 7 chunking expressions
    for col in range(2, min(9, len(en_row))):
        expr_en = en_row[col].strip() if col < len(en_row) and en_row[col] else ''
        expr_kr = kr_row[col].strip() if col < len(kr_row) and kr_row[col] else ''
        
        if expr_en:
            slug = expr_en.replace(' ', '_')
            if expr_kr:
                mapping[slug] = expr_kr
    
    i += 2

# Merge sentence mapping into main mapping
for k, v in sentence_mapping.items():
    if k not in mapping and v:
        mapping[k] = v

# Save as JSON
output_path = r'C:\work\chunking-english\asset\kr_mapping.json'
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(mapping, f, ensure_ascii=False, indent=2)

print(f"Total mappings: {len(mapping)}")
print(f"Saved to: {output_path}")

# Show some samples
print("\n=== Sample Mappings ===")
samples = list(mapping.items())[:20]
for slug, kr in samples:
    print(f"  {slug} -> {kr}")
