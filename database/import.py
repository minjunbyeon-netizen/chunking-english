"""
청킹잉글리시 엑셀 → SQL 임포트 스크립트
------------------------------------------
실행 방법: python database/import.py
결과물:    database/data.sql  (phpMyAdmin에서 임포트)
"""

import openpyxl
import os
import re

# ─── 경로 설정 ────────────────────────────────────────────────
BASE_PATH  = r"C:\xampp\htdocs\03_chunking"
EXCEL_PATH = os.path.join(BASE_PATH, "asset", "청킹 Basic _20260303.xlsx")
IMG_BASE   = os.path.join(BASE_PATH, "asset", "img")
OUTPUT_SQL = os.path.join(BASE_PATH, "database", "data.sql")
MAX_DAY    = 50   # 50일치만 처리
# ──────────────────────────────────────────────────────────────


def is_english(text):
    """영어 텍스트 여부 판단 (한글 없고 알파벳 있으면 영어)"""
    if not text:
        return False
    t = str(text).strip()
    return bool(re.search(r'[a-zA-Z]', t)) and not bool(re.search(r'[가-힣]', t))


def to_snake(text):
    """'have a dream' → 'have_a_dream'"""
    if not text:
        return ''
    return str(text).strip().replace(' ', '_')


def get_image_path(day_number, global_verb_num, verb_en, expression_en):
    """
    이미지 파일 실제 경로 확인 후 상대 경로 반환.
    없으면 None 반환.
    파일명 끝 숫자 오류(be_proud_57.png 등) → 자동 미매핑.
    """
    folder_name = f"day {day_number}"
    verb_folder  = f"{global_verb_num:02d}. {verb_en}"
    filename     = to_snake(expression_en) + ".png"
    full_path    = os.path.join(IMG_BASE, folder_name, verb_folder, filename)

    if os.path.exists(full_path):
        # 상대 경로 (슬래시 통일)
        rel = f"asset/img/day {day_number}/{global_verb_num:02d}. {verb_en}/{filename}"
        return rel
    return None


def escape_sql(val):
    if val is None:
        return "NULL"
    return "'" + str(val).replace("'", "''").replace("\\", "\\\\") + "'"


def main():
    print(f"Loading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["청킹 Baisc"]

    inserts_days        = []
    inserts_verbs       = []
    inserts_expressions = []

    day_number       = 1
    verb_in_day      = 0   # 0~2, 3이 되면 day_number++
    global_verb_num  = 1
    day_id_map       = {}  # day_number → day_id (1-indexed)

    stats = {"mapped": 0, "null_img": 0, "skipped_rows": 0}
    row_idx = 22

    while row_idx <= ws.max_row:
        if day_number > MAX_DAY:
            break

        col_a = ws.cell(row_idx, 1).value
        col_c = ws.cell(row_idx, 3).value

        # ── Day 마커 행 ("Day 5", "Day 10" ...) = 블록 끝 구분선, 스킵 ──
        if col_a and re.match(r'^\s*Day\s+\d+', str(col_a)):
            row_idx += 1
            continue

        # ── 빈 행 ──
        if col_a is None and col_c is None:
            stats["skipped_rows"] += 1
            row_idx += 1
            continue

        # ── 영어 동사 행 ──
        if is_english(col_c):
            en_sentence     = str(col_a).strip() if col_a else ""
            verb_en         = str(col_c).strip()
            expressions_en  = [ws.cell(row_idx, j).value for j in range(7, 14)]

            # 다음 행 = 한국어
            next_row = row_idx + 1
            kr_sentence     = ws.cell(next_row, 1).value
            verb_kr         = ws.cell(next_row, 3).value
            expressions_kr  = [ws.cell(next_row, j).value for j in range(7, 14)]

            kr_sentence = str(kr_sentence).strip() if kr_sentence else ""
            verb_kr     = str(verb_kr).strip()     if verb_kr     else ""

            # Day INSERT (처음 나올 때 한 번만)
            if day_number not in day_id_map:
                day_id = len(day_id_map) + 1
                day_id_map[day_number] = day_id
                inserts_days.append(
                    f"({day_id}, {day_number}, DATE_ADD('2026-01-01', INTERVAL {day_number - 1} DAY), 1)"
                )

            day_id       = day_id_map[day_number]
            verb_id      = global_verb_num
            order_in_day = verb_in_day + 1

            inserts_verbs.append(
                f"({verb_id}, {day_id}, {order_in_day}, {global_verb_num}, "
                f"{escape_sql(verb_en)}, {escape_sql(verb_kr)}, "
                f"{escape_sql(en_sentence)}, {escape_sql(kr_sentence)})"
            )

            # 표현 7개
            for i, (exp_en, exp_kr) in enumerate(zip(expressions_en, expressions_kr)):
                if not exp_en:
                    continue
                exp_en = str(exp_en).strip()
                exp_kr = str(exp_kr).strip() if exp_kr else None

                img_path = get_image_path(day_number, global_verb_num, verb_en, exp_en)
                if img_path:
                    stats["mapped"] += 1
                else:
                    stats["null_img"] += 1

                expr_id = len(inserts_expressions) + 1
                inserts_expressions.append(
                    f"({expr_id}, {verb_id}, {i + 1}, "
                    f"{escape_sql(exp_en)}, {escape_sql(exp_kr)}, "
                    f"{escape_sql(img_path)})"
                )

            verb_in_day     += 1
            global_verb_num += 1

            if verb_in_day >= 3:
                verb_in_day  = 0
                day_number  += 1

            row_idx += 2   # 영어+한국어 쌍 건너뜀
            continue

        stats["skipped_rows"] += 1
        row_idx += 1

    # ── SQL 파일 작성 ──
    print(f"Writing SQL: {OUTPUT_SQL}")
    with open(OUTPUT_SQL, "w", encoding="utf-8") as f:
        f.write("SET NAMES utf8mb4;\n")
        f.write("USE chunking_english;\n\n")
        f.write("SET FOREIGN_KEY_CHECKS = 0;\n")
        f.write("TRUNCATE TABLE expressions;\n")
        f.write("TRUNCATE TABLE verbs;\n")
        f.write("TRUNCATE TABLE days;\n")
        f.write("SET FOREIGN_KEY_CHECKS = 1;\n\n")

        # days
        f.write("-- ① days\n")
        f.write("INSERT INTO days (id, day_number, release_date, is_active) VALUES\n")
        f.write(",\n".join(inserts_days) + ";\n\n")

        # verbs
        f.write("-- ② verbs\n")
        f.write("INSERT INTO verbs (id, day_id, order_num, global_num, verb_en, verb_kr, sentence_en, sentence_kr) VALUES\n")
        f.write(",\n".join(inserts_verbs) + ";\n\n")

        # expressions
        f.write("-- ③ expressions\n")
        f.write("INSERT INTO expressions (id, verb_id, order_num, expression_en, expression_kr, image_path) VALUES\n")
        f.write(",\n".join(inserts_expressions) + ";\n\n")

    print("\nDone!")
    print(f"   Days       : {len(inserts_days)}")
    print(f"   Verbs      : {len(inserts_verbs)}")
    print(f"   Expressions: {len(inserts_expressions)}")
    print(f"   img mapped : {stats['mapped']}")
    print(f"   img NULL   : {stats['null_img']}")
    print(f"   rows skipped: {stats['skipped_rows']}")


if __name__ == "__main__":
    main()
