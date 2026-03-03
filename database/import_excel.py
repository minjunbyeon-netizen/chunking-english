"""
청킹잉글리시 Excel → DB 임포트 스크립트
-----------------------------------------
실행: python database/import_excel.py

동작:
  1. 기존 image_path / audio_path 백업
  2. days / verbs / expressions 전체 재임포트
  3. expression_en 매칭으로 경로 복원
"""

import openpyxl
import pymysql
import sys

EXCEL_PATH = r"C:\Users\USER\Desktop\260303.xlsx"

# ── DB 연결 ───────────────────────────────────────────────────
def get_db():
    return pymysql.connect(
        host="localhost", user="root", password="",
        database="chunking_english", charset="utf8mb4",
        autocommit=False
    )

# ── 구분자 행 판별 ────────────────────────────────────────────
def is_separator(row):
    vals = [v for v in row if v is not None and str(v).strip() not in ('', '-')]
    return len(vals) == 0

# ── Excel 파싱 ────────────────────────────────────────────────
def parse_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    raw = list(ws.iter_rows(min_row=2, values_only=True))

    days = []
    current = []
    i = 0

    while i < len(raw):
        row = raw[i]

        if is_separator(row):
            if current:
                days.append(current)
                current = []
            i += 1
            continue

        en_row = row
        # 다음 행이 한국어 행인지 확인
        if i + 1 < len(raw) and not is_separator(raw[i + 1]):
            kr_row = raw[i + 1]
            i += 2
        else:
            kr_row = [None] * max(len(row), 9)
            i += 1

        def cell(r, idx):
            v = r[idx] if idx < len(r) else None
            return str(v).strip() if v is not None else ''

        verb_en    = cell(en_row, 1)
        verb_kr    = cell(kr_row, 1)
        sent_en    = cell(en_row, 0)
        sent_kr    = cell(kr_row, 0)

        if not verb_en or verb_en == '-':
            i += 0
            continue

        expressions = []
        for j in range(2, 9):
            expr_en = cell(en_row, j)
            expr_kr = cell(kr_row, j)
            if expr_en and expr_en != '-':
                expressions.append({'en': expr_en, 'kr': expr_kr or ''})

        current.append({
            'verb_en':     verb_en.strip(),
            'verb_kr':     verb_kr.strip(),
            'sentence_en': sent_en.strip(),
            'sentence_kr': sent_kr.strip(),
            'expressions': expressions,
        })

    if current:
        days.append(current)

    return days


# ── 메인 ─────────────────────────────────────────────────────
def main():
    print("Excel 파싱 중...")
    days = parse_excel(EXCEL_PATH)
    total_verbs = sum(len(d) for d in days)
    total_exprs = sum(len(v['expressions']) for d in days for v in d)
    print(f"  → {len(days)}일 / {total_verbs}개 동사 / {total_exprs}개 표현\n")

    db  = get_db()
    cur = db.cursor()

    # ── 1. 기존 경로 백업 ─────────────────────────────────────
    print("기존 image_path / audio_path 백업 중...")
    cur.execute("SELECT expression_en, image_path, audio_path FROM expressions WHERE image_path IS NOT NULL OR audio_path IS NOT NULL")
    path_backup = {}
    for row in cur.fetchall():
        key = str(row[0]).strip().lower()
        path_backup[key] = {'image_path': row[1], 'audio_path': row[2]}
    print(f"  → {len(path_backup)}건 백업 완료\n")

    # ── 2. 기존 데이터 삭제 ───────────────────────────────────
    print("기존 데이터 삭제 중...")
    cur.execute("SET FOREIGN_KEY_CHECKS = 0")
    cur.execute("TRUNCATE TABLE expressions")
    cur.execute("TRUNCATE TABLE verbs")
    cur.execute("TRUNCATE TABLE days")
    cur.execute("SET FOREIGN_KEY_CHECKS = 1")
    db.commit()
    print("  → 삭제 완료\n")

    # ── 3. 새 데이터 삽입 ─────────────────────────────────────
    print("새 데이터 삽입 중...")
    global_num = 0
    restored   = 0

    for day_num, verbs in enumerate(days, 1):
        cur.execute(
            "INSERT INTO days (day_number) VALUES (%s)",
            (day_num,)
        )
        day_id = cur.lastrowid

        for v_order, verb in enumerate(verbs, 1):
            global_num += 1
            cur.execute(
                """INSERT INTO verbs
                   (day_id, order_num, global_num, verb_en, verb_kr, sentence_en, sentence_kr)
                   VALUES (%s,%s,%s,%s,%s,%s,%s)""",
                (day_id, v_order, global_num,
                 verb['verb_en'], verb['verb_kr'],
                 verb['sentence_en'], verb['sentence_kr'])
            )
            verb_id = cur.lastrowid

            for e_order, expr in enumerate(verb['expressions'], 1):
                # 경로 복원
                key        = expr['en'].strip().lower()
                image_path = path_backup.get(key, {}).get('image_path')
                audio_path = path_backup.get(key, {}).get('audio_path')
                if image_path or audio_path:
                    restored += 1

                cur.execute(
                    """INSERT INTO expressions
                       (verb_id, order_num, expression_en, expression_kr, image_path, audio_path)
                       VALUES (%s,%s,%s,%s,%s,%s)""",
                    (verb_id, e_order,
                     expr['en'], expr['kr'],
                     image_path, audio_path)
                )

        if day_num % 10 == 0:
            print(f"  Day {day_num} 완료...")

    db.commit()
    db.close()

    print(f"""
==============================================
  임포트 완료!
  Days     : {len(days)}일
  Verbs    : {total_verbs}개
  Exprs    : {total_exprs}개
  경로 복원 : {restored}건
==============================================
""")


if __name__ == "__main__":
    main()
