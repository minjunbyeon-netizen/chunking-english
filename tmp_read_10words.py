import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\work\chunking-english\asset\10words.xlsx', read_only=True, data_only=True)

print(f"=== Sheet Names: {wb.sheetnames} ===")

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== Sheet: {name} ===")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    for i, row in enumerate(ws.iter_rows(max_row=30, values_only=True), 1):
        cells = [str(c).strip() if c is not None else '' for c in row]
        print(f"Row {i:4d}: {' | '.join(cells[:15])}")
    
    if ws.max_row and ws.max_row > 30:
        print(f"  ... ({ws.max_row} rows total)")
        # Show some rows from middle
        print("\n--- Rows 55-65 ---")
        for i, row in enumerate(ws.iter_rows(min_row=55, max_row=65, values_only=True), 55):
            cells = [str(c).strip() if c is not None else '' for c in row]
            print(f"Row {i:4d}: {' | '.join(cells[:15])}")

wb.close()
