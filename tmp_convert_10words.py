import openpyxl, json, re, sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\work\chunking-english\asset\10words.xlsx', read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

data = {}
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    cells = [str(c).strip() if c is not None else '' for c in row]
    if not cells[0]:
        continue
    
    # Day 번호 정규화: "Day1" or "Day 2" → "1" or "2"
    day_raw = cells[0]
    m = re.search(r'(\d+)', day_raw)
    if not m:
        continue
    day_num = str(int(m.group(1)))
    
    entry = {
        'type_ab': cells[1],           # A or B
        'front_type': cells[2],        # 기본 or 변화
        'chunk_type': cells[3],        # to(부정사), ing(동명사), 전치사, 부사절, 등위절
        'back_type': cells[4],         # 변화 or 기본
        'sub_type': cells[5],          # ing(동명사) etc
        'en_front': cells[6],          # English front
        'en_back': cells[7],           # English back
        'kr_front': cells[10] if len(cells) > 10 else '',  # Korean front
        'kr_back': cells[11] if len(cells) > 11 else '',   # Korean back
    }
    
    if day_num not in data:
        data[day_num] = []
    data[day_num].append(entry)

wb.close()

out_path = r'C:\work\chunking-english\asset\10words_mapping.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Total days: {len(data)}")
for d in ['1', '2', '3']:
    if d in data:
        print(f"Day {d}: {len(data[d])} sentences")
        for j, s in enumerate(data[d]):
            print(f"  [{j+1}] {s['en_front']} {s['en_back']} | {s['kr_front']} {s['kr_back']} | type={s['chunk_type']}")
print(f"\nSaved to: {out_path}")
